from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from fastapi import FastAPI
from pandasql import sqldf
import xlsx_streaming
import pandas as pd
import uvicorn
import random
import io


# Initialize FastAPI app
app = FastAPI()


# Dummy dataframe
gen_data = [[random.random() for _ in range(3)] for _ in range(15)]

df = pd.DataFrame(gen_data, columns=['un', 'deux', 'trois'])


# Select the dataframe with SQL to imitate a Databae request
def select_table(offset):
    truc = sqldf(f'''
        SELECT
            *
        FROM
            df
        ORDER BY
            un
        LIMIT
            5
        OFFSET
            {offset}
        ''')
    return truc


# Generate header and first row for xlsx_streaming data typing
async def generate_header_buffer(data, sheet_nbr, book):

    # request data as whole dataframe
    snow_data = data
    headers = snow_data.columns
    # book = Workbook()

    book.create_sheet(title=f"sheet{sheet_nbr}")

    if 'Sheet' in book.sheetnames:
        sheet_to_remove = book['Sheet']
        book.remove(sheet_to_remove)

    book.active = sheet_nbr
    sheet = book.active

    sheet.append(list(headers))
    sheet.append(list(snow_data.iloc[0]))

    buffer = io.BytesIO()
    book.save(buffer)

    return buffer


# Stream the data in batches
async def generate_excel(data_len):

    book = Workbook()

    for sheet in range(data_len // 5):

        offset = sheet * 5

        data = select_table(offset)

        buffer = await generate_header_buffer(data, sheet, book)

        for batch in xlsx_streaming.stream_queryset_as_xlsx(qs=data.itertuples(index=False), xlsx_template=buffer, batch_size=1):
            yield batch


# -----------------------------------------------------------------------------------------------------


# Using StreamingResponse to stream the data
@app.get("/extract", response_class=StreamingResponse)
async def extract(filename: str):

    data_len = len(df)

    return StreamingResponse(
        content=generate_excel(data_len),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            "Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )


# run the app locally
if __name__ == '__main__':
    uvicorn.run(app, host='127.0.0.1', port=8000)
